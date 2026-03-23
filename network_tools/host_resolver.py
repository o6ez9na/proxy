import socket
from ipwhois import IPWhois
import pandas as pd
import ipaddress
import requests
import base64
import re
import sys
import urllib.parse
from urllib.parse import urlparse


def extract_data(vless_text: str):
    results = []

    for line in vless_text.splitlines():
        if not line.startswith("vless://"):
            continue

        domain_match = re.search(r"@([^:]+):", line)
        domain = domain_match.group(1) if domain_match else "UNKNOWN"

        name = ""
        if "#" in line:
            encoded_name = line.split("#", 1)[1]
            name = urllib.parse.unquote(encoded_name)

        results.append((domain, name))

    unique = {}
    for domain, name in results:
        if domain not in unique:
            unique[domain] = name

    return list(unique.items())


from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def resolve(data, output_file):
    results = []

    for domain, name in data:
        try:
            try:
                ipaddress.ip_address(domain)
                ip = domain
            except ValueError:
                ip = socket.gethostbyname(domain)

            try:
                obj = IPWhois(ip)
                res = obj.lookup_rdap()
                provider = (
                    res.get("network", {}).get("name")
                    or res.get("network", {}).get("org")
                    or "Unknown"
                )
            except Exception:
                provider = "WHOIS ERROR"

            print(f"{domain} - {ip} - {provider} - {name}")
            results.append([domain, ip, provider, name])

        except Exception as e:
            print(f"{domain} - ERROR ({e})")
            results.append([domain, "ERROR", str(e), name])

    df = pd.DataFrame(results, columns=["Domain", "IP", "Provider", "Name"])
    df.to_excel(output_file, index=False)

    wb = load_workbook(output_file)
    ws = wb.active

    fill_antiglush = PatternFill(
        start_color="496B74", end_color="496B74", fill_type="solid"
    )
    fill_white = PatternFill(
        start_color="D8DB9A", end_color="D8DB9A", fill_type="solid"
    )

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        name_cell = row[3]

        if name_cell.value:
            if "Антиглушилка" in name_cell.value or "LTE" in name_cell.value:
                for cell in row:
                    cell.fill = fill_antiglush

            elif "Белые" in name_cell.value:
                for cell in row:
                    cell.fill = fill_white

    wb.save(output_file)

    print(f"\nРезультат сохранен в {output_file}")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Использование: python script.py <URL>")
        sys.exit(1)

    url = sys.argv[1]

    try:
        response = requests.get(url, timeout=10)
        response.raise_for_status()
    except Exception as e:
        print(f"Ошибка запроса: {e}")
        sys.exit(1)

    try:
        decoded = base64.b64decode(response.text).decode("utf-8")
    except Exception as e:
        print(f"Ошибка декодирования base64: {e}")
        sys.exit(1)

    data = extract_data(decoded)

    print(f"Найдено доменов: {len(data)}\n")

    parsed = urlparse(url)
    filename_base = parsed.netloc.replace(":", "_")
    output_file = f"../{filename_base}.xlsx"

    resolve(data, output_file)
